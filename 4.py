
import time, random, requests
import pandas as pd
import numpy as np
from datetime import datetime
import streamlit as st
from selenium import webdriver
from selenium.webdriver.firefox.options import Options
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from sklearn.linear_model import LinearRegression, Ridge, Lasso
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, r2_score
import os
import plotly.express as px
import warnings
warnings.filterwarnings('ignore')

# ================= CONFIG =================
SAVE_FOLDER = r"C:\Users\RAGHURAM MAHATA\Desktop\NSE_STOCK"
INDICES = ["NIFTY", "BANKNIFTY", "FINNIFTY", "MIDCPNIFTY"]
os.makedirs(SAVE_FOLDER, exist_ok=True)

# ---------------- UTILITY FUNCTIONS ----------------
@st.cache_data(ttl=300)
def get_cookies_from_firefox():
    options = Options()
    options.add_argument("--headless")
    driver = webdriver.Firefox(options=options)
    st.info("Opening NSE page in Firefox to fetch cookies...")
    driver.get("https://www.nseindia.com/option-chain")
    time.sleep(10)
    cookies = driver.get_cookies()
    driver.quit()
    return "; ".join([f"{c['name']}={c['value']}" for c in cookies])

def fetch_option_chain(symbol, cookie_string):
    url = f"https://www.nseindia.com/api/option-chain-indices?symbol={symbol}" if symbol in INDICES else f"https://www.nseindia.com/api/option-chain-equities?symbol={symbol}"
    headers = {"User-Agent": "Mozilla/5.0", "Accept": "application/json, text/plain, */*", "Referer": "https://www.nseindia.com/option-chain", "Cookie": cookie_string}
    for attempt in range(3):
        try:
            resp = requests.get(url, headers=headers, timeout=30)
            resp.raise_for_status()
            return resp.json()
        except Exception as e:
            st.error(f"Attempt {attempt+1} failed for {symbol}: {e}")
            time.sleep(random.randint(3, 7))
    return None

def parse_data(symbol, data):
    if not data: return pd.DataFrame()
    expiry_dates = data.get("records", {}).get("expiryDates", [])
    if not expiry_dates: return pd.DataFrame()
    expiry = expiry_dates[0]
    
    records = []
    for item in data.get("records", {}).get("data", []):
        if item.get("expiryDate") != expiry: continue
        ce, pe = item.get("CE", {}), item.get("PE", {})
        records.append({
            "STRIKE": item["strikePrice"],
            "CALL_OI": ce.get("openInterest", 0),
            "CALL_CHNG_IN_OI": ce.get("changeinOpenInterest", 0),
            "CALL_IV": ce.get("impliedVolatility", 0),
            "CALL_LTP": ce.get("lastPrice", 0),
            "PUT_OI": pe.get("openInterest", 0),
            "PUT_CHNG_IN_OI": pe.get("changeinOpenInterest", 0),
            "PUT_IV": pe.get("impliedVolatility", 0),
            "PUT_LTP": pe.get("lastPrice", 0)
        })
    
    df = pd.DataFrame(records)
    return df.sort_values("STRIKE").reset_index(drop=True) if not df.empty else df

# ---------------- ADVANCED ANALYTICS ----------------
def calculate_analytics(df):
    if df.empty: return {}
    
    # Calculate total OI and deltas
    df['TOTAL_OI'] = df['CALL_OI'] + df['PUT_OI']
    df['OI_RATIO'] = df['PUT_OI'] / df['CALL_OI'].replace(0, 1)
    df['DELTA_CALL'] = df['CALL_OI'] / df['TOTAL_OI'].replace(0, 1)
    df['DELTA_PUT'] = df['PUT_OI'] / df['TOTAL_OI'].replace(0, 1)
    df['IV_DIFF'] = df['CALL_IV'] - df['PUT_IV']
    
    # Basic metrics
    total_put_oi = df['PUT_OI'].sum()
    total_call_oi = df['CALL_OI'].sum()
    pcr = total_put_oi / total_call_oi if total_call_oi > 0 else 0
    
    # Sentiment classification
    if pcr > 1.5:
        sentiment = "Strongly Bullish"
        sentiment_score = 2
    elif pcr > 1.2:
        sentiment = "Bullish"
        sentiment_score = 1
    elif pcr < 0.5:
        sentiment = "Strongly Bearish"
        sentiment_score = -2
    elif pcr < 0.8:
        sentiment = "Bearish"
        sentiment_score = -1
    else:
        sentiment = "Neutral"
        sentiment_score = 0
    
    # Key levels
    strongest_support = df.loc[df['PUT_OI'].idxmax(), 'STRIKE']
    strongest_resistance = df.loc[df['CALL_OI'].idxmax(), 'STRIKE']
    max_pain = df.loc[df['TOTAL_OI'].idxmin(), 'STRIKE']
    
    # Top values
    top_3_call = df.nlargest(3, 'CALL_OI')[['STRIKE', 'CALL_OI']]
    top_3_put = df.nlargest(3, 'PUT_OI')[['STRIKE', 'PUT_OI']]
    top_3_call_change = df.nlargest(3, 'CALL_CHNG_IN_OI')[['STRIKE', 'CALL_CHNG_IN_OI']]
    top_3_put_change = df.nlargest(3, 'PUT_CHNG_IN_OI')[['STRIKE', 'PUT_CHNG_IN_OI']]
    top3_call_iv = df.nlargest(3, 'CALL_IV')[['STRIKE', 'CALL_IV']]
    top3_put_iv = df.nlargest(3, 'PUT_IV')[['STRIKE', 'PUT_IV']]
    
    # Predicted values
    predicted_max_pain = np.average(df['STRIKE'], weights=df['TOTAL_OI'])
    predicted_range = (predicted_max_pain - df['STRIKE'].std(), predicted_max_pain + df['STRIKE'].std())
    
    # Volume analysis
    total_call_volume = df['CALL_CHNG_IN_OI'].sum()
    total_put_volume = df['PUT_CHNG_IN_OI'].sum()
    volume_ratio = total_put_volume / total_call_volume if total_call_volume > 0 else 0
    
    return {
        "pcr": pcr,
        "sentiment": sentiment,
        "sentiment_score": sentiment_score,
        "strongest_support": strongest_support,
        "strongest_resistance": strongest_resistance,
        "max_pain": max_pain,
        "predicted_max_pain": predicted_max_pain,
        "predicted_range": predicted_range,
        "top_3_call": top_3_call,
        "top_3_put": top_3_put,
        "top_3_call_change": top_3_call_change,
        "top_3_put_change": top_3_put_change,
        "top3_call_iv": top3_call_iv,
        "top3_put_iv": top3_put_iv,
        "volume_ratio": volume_ratio,
        "delta_table": df[['STRIKE', 'DELTA_CALL', 'DELTA_PUT']],
        "df": df
    }

# ---------------- MACHINE LEARNING MODELS ----------------
def train_ml_models(df):
    """Train multiple ML models for prediction"""
    if df.empty or len(df) < 10:
        return {}, [], []
    
    # Prepare features and target
    X = df[['CALL_OI', 'PUT_OI', 'CALL_CHNG_IN_OI', 'PUT_CHNG_IN_OI', 
            'CALL_IV', 'PUT_IV', 'CALL_LTP', 'PUT_LTP']].fillna(0)
    y = df['STRIKE']
    
    # Split data
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)
    
    # Scale features
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)
    
    # Initialize models
    models = {
        'Linear Regression': LinearRegression(),
        'Ridge Regression': Ridge(alpha=1.0),
        'Lasso Regression': Lasso(alpha=0.1),
        'Random Forest': RandomForestRegressor(n_estimators=100, random_state=42),
        'Gradient Boosting': GradientBoostingRegressor(n_estimators=100, random_state=42)
    }
    
    # Train and evaluate models
    results = {}
    for name, model in models.items():
        model.fit(X_train_scaled, y_train)
        y_pred = model.predict(X_test_scaled)
        mae = mean_absolute_error(y_test, y_pred)
        r2 = r2_score(y_test, y_pred)
        results[name] = {'model': model, 'mae': mae, 'r2': r2, 'scaler': scaler}
    
    # Get best model
    best_model_name = min(results, key=lambda x: results[x]['mae'])
    best_model = results[best_model_name]['model']
    best_scaler = results[best_model_name]['scaler']
    
    # Make predictions on full dataset
    X_full_scaled = best_scaler.transform(X)
    df['ML_PREDICTED_STRIKE'] = best_model.predict(X_full_scaled)
    
    # Find top calls and puts based on ML predictions
    top_calls = df.nlargest(3, 'ML_PREDICTED_STRIKE')['STRIKE'].tolist()
    top_puts = df.nsmallest(3, 'ML_PREDICTED_STRIKE')['STRIKE'].tolist()
    
    return results, top_calls, top_puts

# ---------------- ADVANCED CHARTS ----------------
def create_oi_chart(df):
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=df['STRIKE'], y=df['CALL_OI'], mode='lines', name='Call OI',
                             line=dict(shape='spline', smoothing=1.3, color='red')))
    fig.add_trace(go.Scatter(x=df['STRIKE'], y=df['PUT_OI'], mode='lines', name='Put OI',
                             line=dict(shape='spline', smoothing=1.3, color='green')))
    fig.update_layout(
        title="Open Interest Distribution", 
        xaxis_title="Strike Price",
        yaxis_title="Open Interest", 
        height=250, 
        margin=dict(t=30, b=10, l=10, r=10),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    return fig

def create_sentiment_chart(df):
    df['SENTIMENT'] = df['CALL_OI'] - df['PUT_OI']
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=df['STRIKE'], 
        y=df['SENTIMENT'],
        marker_color=['green' if val > 0 else 'red' for val in df['SENTIMENT']]
    ))
    fig.update_layout(
        title="Sentiment (Call OI - Put OI)", 
        xaxis_title="Strike", 
        yaxis_title="Call-Put OI",
        height=250, 
        margin=dict(t=30, b=10, l=10, r=10)
    )
    return fig

def create_iv_comparison_chart(df):
    fig = go.Figure()
    fig.add_trace(go.Scatter(
        x=df['STRIKE'], 
        y=df['CALL_IV'], 
        mode='lines', 
        name='Call IV',
        line=dict(color='blue', shape='spline', smoothing=1.3)
    ))
    fig.add_trace(go.Scatter(
        x=df['STRIKE'], 
        y=df['PUT_IV'], 
        mode='lines', 
        name='Put IV',
        line=dict(color='red', shape='spline', smoothing=1.3)
    ))
    fig.update_layout(
        title="Implied Volatility Comparison", 
        xaxis_title="Strike Price",
        yaxis_title="IV (%)", 
        height=250, 
        margin=dict(t=30, b=10, l=10, r=10),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    return fig

def create_ml_prediction_chart(df, analytics, top_calls, top_puts):
    fig = go.Figure()
    
    # Add actual strikes
    fig.add_trace(go.Scatter(
        x=df['STRIKE'], 
        y=df['TOTAL_OI'], 
        mode='lines+markers', 
        name='Total OI',
        line=dict(color='blue', shape='spline', smoothing=1.3)
    ))
    
    # Add ML predicted strikes
    fig.add_trace(go.Scatter(
        x=df['ML_PREDICTED_STRIKE'], 
        y=df['TOTAL_OI'], 
        mode='markers', 
        name='ML Predicted',
        marker=dict(color='orange', size=8, symbol='diamond')
    ))
    
    # Add max pain
    fig.add_vline(
        x=analytics['max_pain'], 
        line_dash="dash", 
        line_color="purple", 
        annotation_text="Max Pain"
    )
    
    # Add top calls
    for strike in top_calls:
        fig.add_vline(
            x=strike, 
            line_dash="dot", 
            line_color="green", 
            annotation_text=f"C {strike}"
        )
    
    # Add top puts
    for strike in top_puts:
        fig.add_vline(
            x=strike, 
            line_dash="dot", 
            line_color="red", 
            annotation_text=f"P {strike}"
        )
    
    fig.update_layout(
        title="ML Predictions & Key Levels", 
        xaxis_title="Strike Price",
        yaxis_title="Total OI", 
        height=300, 
        margin=dict(t=30, b=10, l=10, r=10),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
    )
    return fig

def create_model_performance_chart(ml_results):
    models = list(ml_results.keys())
    mae_scores = [ml_results[model]['mae'] for model in models]
    r2_scores = [ml_results[model]['r2'] for model in models]
    
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=models, 
        y=mae_scores, 
        name='MAE',
        marker_color='lightblue'
    ))
    fig.add_trace(go.Bar(
        x=models, 
        y=r2_scores, 
        name='R²',
        marker_color='orange',
        yaxis='y2'
    ))
    
    fig.update_layout(
        title="ML Model Performance Comparison",
        xaxis_title="Models",
        yaxis_title="MAE (Lower is better)",
        yaxis2=dict(title="R² (Higher is better)", overlaying='y', side='right'),
        barmode='group',
        height=300,
        margin=dict(t=30, b=10, l=10, r=10)
    )
    return fig

# ---------------- EXCEL EXPORT ----------------
def save_to_excel(df, analytics, symbol, ml_results, top_calls, top_puts):
    wb = Workbook()
    
    # Option Chain sheet
    ws1 = wb.active
    ws1.title = "OptionChain"
    
    # Formatting styles
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Write option chain data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=value)
            cell.alignment = center_align
            cell.border = thin_border
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = bold_font
    
    # Analytics sheet
    ws2 = wb.create_sheet("Analytics")
    analytics_data = [
        ["PCR", analytics["pcr"]],
        ["Sentiment", analytics["sentiment"]],
        ["Strongest Support", analytics["strongest_support"]],
        ["Strongest Resistance", analytics["strongest_resistance"]],
        ["Max Pain", analytics["max_pain"]],
        ["Predicted Max Pain", analytics["predicted_max_pain"]],
        ["Volume Ratio (P/C)", analytics["volume_ratio"]],
        ["ML Top Calls", ", ".join(map(str, top_calls))],
        ["ML Top Puts", ", ".join(map(str, top_puts))]
    ]
    
    for i, (label, value) in enumerate(analytics_data, 1):
        ws2.cell(row=i, column=1, value=label).font = bold_font
        ws2.cell(row=i, column=2, value=value)
    
    # ML Results sheet
    ws3 = wb.create_sheet("ML_Results")
    ws3.append(["Model", "MAE", "R²"])
    for model, results in ml_results.items():
        ws3.append([model, results['mae'], results['r2']])
    
    # Save file
    file_path = os.path.join(SAVE_FOLDER, f"{symbol}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(file_path)
    return file_path

# ---------------- Styled Option Chain Table ----------------
def styled_option_chain(df):
    max_call = df['CALL_OI'].max()
    max_put = df['PUT_OI'].max()

    def color_rows(row):
        if row['CALL_OI'] > row['PUT_OI']:
            return ['background-color: #4CAF50; color: white;' if col not in ['STRIKE'] else 'font-weight: bold;' for col in row.index]
        elif row['PUT_OI'] > row['CALL_OI']:
            return ['background-color: #f44336; color: white;' if col not in ['STRIKE'] else 'font-weight: bold;' for col in row.index]
        else:
            return ['background-color: #FFEB3B;' if col not in ['STRIKE'] else 'font-weight: bold;' for col in row.index]

    def bar_call(val):
        width = int((val / max_call) * 100) if max_call > 0 else 0
        return f'background: linear-gradient(to right, #4CAF50 {width}%, transparent 0%);'

    def bar_put(val):
        width = int((val / max_put) * 100) if max_put > 0 else 0
        return f'background: linear-gradient(to right, #f44336 {width}%, transparent 0%);'

    styler = df.style.apply(color_rows, axis=1)\
                     .set_properties(subset=['CALL_OI'], **{'background': df['CALL_OI'].apply(bar_call)})\
                     .set_properties(subset=['PUT_OI'], **{'background': df['PUT_OI'].apply(bar_put)})\
                     .set_properties(subset=['STRIKE'], **{'font-weight': 'bold'})\
                     .set_table_styles([{'selector':'th','props':[('font-weight','bold'), ('background-color','#1976D2'),('color','white')]}])\
                     .format({col: "{:,.0f}" for col in df.select_dtypes(include=np.number).columns})
    return styler

# ================= STREAMLIT APP =================
def run_streamlit_app():
    st.set_page_config(page_title="NSE Option Chain Dashboard", layout="wide", initial_sidebar_state="expanded")

    # ---------------- Professional CSS ----------------
    st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 1rem;
        font-weight: 700;
        background: linear-gradient(135deg, #1f77b4, #ff7f0e);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        padding: 10px;
        border-radius: 10px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
    }
    .metric-card {
        background: linear-gradient(135deg, #f8f9fa, #e9ecef);
        border-radius: 10px;
        padding: 15px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        text-align: center;
        margin-bottom: 15px;
        border-left: 4px solid #1f77b4;
        transition: transform 0.3s ease;
    }
    .metric-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 6px 12px rgba(0, 0, 0, 0.15);
    }
    .metric-title {
        font-size: 0.9rem;
        color: #6c757d;
        font-weight: 600;
        margin-bottom: 5px;
    }
    .metric-value {
        font-size: 1.4rem;
        color: #212529;
        font-weight: 700;
    }
    .chart-container {
        background: white;
        border-radius: 10px;
        padding: 15px;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
        margin-bottom: 20px;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        white-space: pre-wrap;
        background-color: #f8f9fa;
        border-radius: 8px 8px 0px 0px;
        gap: 8px;
        padding-top: 10px;
        padding-bottom: 10px;
        font-weight: 600;
    }
    .stTabs [aria-selected="true"] {
        background-color: #1f77b4;
        color: white;
    }
    .section-header {
        font-size: 1.4rem;
        color: #1f77b4;
        margin: 20px 0 15px 0;
        font-weight: 600;
        border-bottom: 2px solid #1f77b4;
        padding-bottom: 8px;
    }
    </style>
    """, unsafe_allow_html=True)

    # ---------------- Header ----------------
    st.markdown('<div class="main-header">📊 Advanced NSE Option Chain Analyzer</div>', unsafe_allow_html=True)
    
    # ---------------- Sidebar ----------------
    with st.sidebar:
        st.header("🔧 Configuration")
        symbol = st.text_input("Symbol", "NIFTY", help="Enter NIFTY, BANKNIFTY, or any stock symbol").upper()
        
        st.subheader("ML Settings")
        ml_enabled = st.checkbox("Enable ML Predictions", value=True)
        auto_refresh = st.checkbox("Auto-Refresh (30s)", value=False)
        
        st.subheader("Chart Options")
        show_oi_chart = st.checkbox("Show OI Chart", value=True)
        show_sentiment_chart = st.checkbox("Show Sentiment Chart", value=True)
        show_iv_chart = st.checkbox("Show IV Chart", value=True)
        
        st.info("ℹ️ Data is fetched in real-time from NSE India")
    
    # ---------------- Main Content ----------------
    if auto_refresh:
        time.sleep(30)
        st.rerun()
    
    fetch_clicked = st.button("🚀 Fetch & Analyze Data", type="primary", use_container_width=True)
    
    if fetch_clicked:
        with st.spinner("🔄 Fetching real-time data from NSE..."):
            cookie_string = get_cookies_from_firefox()
            data = fetch_option_chain(symbol, cookie_string)

        if data:
            df = parse_data(symbol, data)
            if not df.empty:
                analytics = calculate_analytics(df)
                
                # Machine Learning
                if ml_enabled:
                    with st.spinner("🤖 Training ML models..."):
                        ml_results, top_calls, top_puts = train_ml_models(df)
                else:
                    ml_results, top_calls, top_puts = {}, [], []
                
                # ---------------- Key Metrics ----------------
                st.markdown('<div class="section-header">📈 Key Metrics</div>', unsafe_allow_html=True)
                
                cols = st.columns(4)
                metrics = [
                    ("PCR", f"{analytics['pcr']:.2f}", "#4CAF50" if analytics['pcr'] > 1 else "#f44336"),
                    ("Sentiment", analytics["sentiment"], "#4CAF50" if analytics['sentiment_score'] > 0 else "#f44336" if analytics['sentiment_score'] < 0 else "#FFC107"),
                    ("Support", analytics["strongest_support"], "#4CAF50"),
                    ("Resistance", analytics["strongest_resistance"], "#f44336")
                ]
                
                for col, (label, value, color) in zip(cols, metrics):
                    with col:
                        st.markdown(f"""
                        <div class="metric-card" style="border-left-color: {color}">
                            <div class="metric-title">{label}</div>
                            <div class="metric-value">{value}</div>
                        </div>
                        """, unsafe_allow_html=True)
                
                # ---------------- ML Predictions ----------------
                if ml_enabled and ml_results:
                    st.markdown('<div class="section-header">🤖 Machine Learning Insights</div>', unsafe_allow_html=True)
                    
                    ml_col1, ml_col2, ml_col3 = st.columns(3)
                    
                    with ml_col1:
                        best_model = min(ml_results, key=lambda x: ml_results[x]['mae'])
                        st.markdown(f"""
                        <div class="metric-card" style="border-left-color: #9C27B0">
                            <div class="metric-title">Best Model</div>
                            <div class="metric-value">{best_model}</div>
                            <div style="font-size: 0.8rem; color: #6c757d;">
                                MAE: {ml_results[best_model]['mae']:.2f}<br>
                                R²: {ml_results[best_model]['r2']:.3f}
                            </div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with ml_col2:
                        calls_str = ", ".join(map(str, top_calls)) if top_calls else "N/A"
                        st.markdown(f"""
                        <div class="metric-card" style="border-left-color: #4CAF50">
                            <div class="metric-title">ML Recommended Calls</div>
                            <div class="metric-value">{calls_str}</div>
                        </div>
                        """, unsafe_allow_html=True)
                    
                    with ml_col3:
                        puts_str = ", ".join(map(str, top_puts)) if top_puts else "N/A"
                        st.markdown(f"""
                        <div class="metric-card" style="border-left-color: #f44336">
                            <div class="metric-title">ML Recommended Puts</div>
                            <div class="metric-value">{puts_str}</div>
                        </div>
                        """, unsafe_allow_html=True)
                
                # ---------------- Charts ----------------
                st.markdown('<div class="section-header">📊 Advanced Charts</div>', unsafe_allow_html=True)
                
                tab1, tab2, tab3, tab4 = st.tabs(["OI Analysis", "Sentiment Analysis", "IV Analysis", "ML Analysis"])
                
                with tab1:
                    if show_oi_chart:
                        st.plotly_chart(create_oi_chart(df), use_container_width=True)
                
                with tab2:
                    if show_sentiment_chart:
                        st.plotly_chart(create_sentiment_chart(df), use_container_width=True)
                
                with tab3:
                    if show_iv_chart:
                        st.plotly_chart(create_iv_comparison_chart(df), use_container_width=True)
                
                with tab4:
                    if ml_enabled and ml_results:
                        col1, col2 = st.columns(2)
                        with col1:
                            st.plotly_chart(create_ml_prediction_chart(df, analytics, top_calls, top_puts), use_container_width=True)
                        with col2:
                            st.plotly_chart(create_model_performance_chart(ml_results), use_container_width=True)
                
                # ---------------- Option Chain Table ----------------
                st.markdown('<div class="section-header">📋 Option Chain Data</div>', unsafe_allow_html=True)
                st.dataframe(styled_option_chain(df), use_container_width=True, height=400)
                
                # ---------------- Top 5 Calls & Puts ----------------
                st.markdown('<div class="section-header">🔥 Top 5 Calls & Puts by Open Interest</div>', unsafe_allow_html=True)
                
                top_5_calls = analytics['df'].nlargest(5, 'CALL_OI')[['STRIKE', 'CALL_OI', 'CALL_IV', 'CALL_LTP']]
                top_5_puts = analytics['df'].nlargest(5, 'PUT_OI')[['STRIKE', 'PUT_OI', 'PUT_IV', 'PUT_LTP']]
                
                col_calls, col_puts = st.columns(2)
                with col_calls:
                    st.write("📈 Top 5 Calls")
                    st.dataframe(top_5_calls, use_container_width=True, height=300)
                with col_puts:
                    st.write("📉 Top 5 Puts")
                    st.dataframe(top_5_puts, use_container_width=True, height=300)
                
                # ---------------- Actionable Insights ----------------
                st.markdown('<div class="section-header">💡 Actionable Insights</div>', unsafe_allow_html=True)
                
                insights = []
                pcr = analytics.get("pcr", 0)
                sentiment = analytics.get("sentiment", "Neutral")
                max_pain = analytics.get("max_pain", None)
                support = analytics.get("strongest_support", None)
                resistance = analytics.get("strongest_resistance", None)
                volume_ratio = analytics.get("volume_ratio", 0)
                
                if pcr > 1.5:
                    insights.append("📊 High PCR indicates bullish market sentiment.")
                elif pcr < 0.5:
                    insights.append("📉 Low PCR indicates bearish market sentiment.")
                else:
                    insights.append("⚖️ PCR near neutral - market indecision.")
                
                if volume_ratio > 1:
                    insights.append("🔼 Put volume > Call volume suggests bearish pressure.")
                else:
                    insights.append("🔽 Call volume > Put volume suggests bullish pressure.")
                
                if max_pain is not None:
                    insights.append(f"🔧 Max Pain at strike {max_pain} could act as price magnet near expiry.")
                
                if support is not None and resistance is not None:
                    insights.append(f"🛑 Support: {support}, Resistance: {resistance} levels are key.")
                
                for insight in insights:
                    st.markdown(f"- {insight}")
                
                # ---------------- Export Data ----------------
                st.markdown('<div class="section-header">💾 Export Results</div>', unsafe_allow_html=True)
                file_path = save_to_excel(df, analytics, symbol, ml_results, top_calls, top_puts)
                st.success(f"✅ Analysis saved to: {file_path}")
                
                with open(file_path, "rb") as f:
                    st.download_button(
                        label="📥 Download Full Report (Excel)",
                        data=f,
                        file_name=os.path.basename(file_path),
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            else:
                st.warning("⚠️ No data found for this symbol. Please check the symbol and try again.")
        else:
            st.error("❌ Failed to fetch option chain data. This might be due to network issues or NSE API changes.")

    # ---------------- Footer ----------------
    st.markdown("---")
    st.caption(f"🕒 Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | "
               f"🤖 ML Models: Linear Regression, Ridge, Lasso, Random Forest, Gradient Boosting")


if __name__ == "__main__":
    run_streamlit_app()
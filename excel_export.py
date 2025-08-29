import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

SAVE_FOLDER = r"C:\Users\RAGHURAM MAHATA\Desktop\NSE_STOCK"
os.makedirs(SAVE_FOLDER, exist_ok=True)

def save_to_excel(df, analytics, symbol, ml_results=None, top_calls=None, top_puts=None):
    """
    Save Option Chain data, analytics, and ML results into a styled Excel report.
    
    Returns:
        str: Path to the saved Excel file
    """
    if df.empty:
        return None

    file_path = os.path.join(
        SAVE_FOLDER, f"{symbol}_OptionChain_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Option Chain Data"

    # ================== Write Option Chain Data ================== #
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Styling
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Borders
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border

    # ================== Add Analytics Sheet ================== #
    ws2 = wb.create_sheet("Analytics")
    ws2.append(["Metric", "Value"])

    for key, val in analytics.items():
        if key == "df":  # skip raw DataFrame
            continue
        ws2.append([key, val])

    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)

    # ================== Add ML Results (if available) ================== #
    if ml_results:
        ws3 = wb.create_sheet("ML_Results")
        ws3.append(["Model", "Accuracy"])
        for model, details in ml_results.items():
            ws3.append([model, details.get("accuracy", 0)])

    # ================== Add Top Calls/Puts (if available) ================== #
    if top_calls and top_puts:
        ws4 = wb.create_sheet("Top Calls & Puts")
        ws4.append(["Top Calls"])
        ws4.append(["Strike", "OI", "IV", "LTP"])
        for rec in top_calls:
            ws4.append(list(rec.values()))

        ws4.append([])
        ws4.append(["Top Puts"])
        ws4.append(["Strike", "OI", "IV", "LTP"])
        for rec in top_puts:
            ws4.append(list(rec.values()))

    wb.save(file_path)
    return file_path
